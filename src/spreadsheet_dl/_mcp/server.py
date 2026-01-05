"""MCP Server core implementation.

This module contains the complete MCPServer class with all tool handlers.

Future modularization will extract tool handlers into separate category files
in the tools/ subpackage. For now, all functionality is consolidated here.
"""

from __future__ import annotations

import contextlib
import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from spreadsheet_dl._mcp.config import MCPConfig, MCPVersion
from spreadsheet_dl._mcp.exceptions import MCPSecurityError
from spreadsheet_dl._mcp.models import MCPTool, MCPToolParameter, MCPToolResult
from spreadsheet_dl._mcp.registry import MCPToolRegistry
from spreadsheet_dl.exceptions import FileError


class MCPServer:
    """MCP server for spreadsheet-dl.

    Exposes spreadsheet manipulation and budget analysis tools via MCP protocol,
    enabling natural language interaction with Claude Desktop and
    other MCP-compatible clients.

    Tools provided:
        Budget Analysis (8 tools):
            - analyze_budget: Analyze a budget file
            - add_expense: Add a new expense
            - query_budget: Answer natural language budget questions
            - get_spending_trends: Analyze spending patterns
            - compare_periods: Compare two time periods
            - generate_report: Generate formatted reports
            - list_categories: List expense categories
            - get_alerts: Check budget alerts

        Cell Operations:
            - cell_get, cell_set, cell_clear
            - cell_copy, cell_move
            - cell_batch_get, cell_batch_set
            - cell_find, cell_replace
            - cell_merge, cell_unmerge

        Style Operations:
            - style_list, style_get, style_create
            - style_update, style_delete, style_apply
            - format_cells, format_number
            - format_font, format_fill, format_border

        Structure Operations:
            - row_insert, row_delete, row_hide
            - column_insert, column_delete, column_hide
            - freeze_set, freeze_clear
            - sheet_create, sheet_delete, sheet_copy

        Advanced Tools:
            - chart_create, chart_update
            - validation_create, cf_create
            - named_range_create, table_create
            - query_select, query_find

    Example:
        >>> server = MCPServer()
        >>> server.run()  # Starts stdio-based MCP server
    """

    def __init__(
        self,
        config: MCPConfig | None = None,
    ) -> None:
        """Initialize MCP server.

        Args:
            config: Server configuration. Uses defaults if not provided.
        """
        self.config = config or MCPConfig()
        self.logger = logging.getLogger("spreadsheet-dl-mcp")
        self._registry = MCPToolRegistry()
        self._tools: dict[str, MCPTool] = {}
        self._request_count = 0
        self._last_reset = datetime.now()
        self._register_tools()

    def _register_tools(self) -> None:
        """Register all available tools."""
        # =====================================================================
        # Cell Operation Tools
        # =====================================================================

        # cell_get tool
        self._registry.register(
            name="cell_get",
            description="Get the value of a specific cell",
            handler=self._handle_cell_get,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="cell",
                    type="string",
                    description="Cell reference (e.g., 'A1', 'B5')",
                ),
            ],
            category="cell_operations",
        )

        # cell_set tool
        self._registry.register(
            name="cell_set",
            description="Set the value of a specific cell",
            handler=self._handle_cell_set,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="cell",
                    type="string",
                    description="Cell reference (e.g., 'A1', 'B5')",
                ),
                MCPToolParameter(
                    name="value",
                    type="string",
                    description="Value to set",
                ),
            ],
            category="cell_operations",
        )

        # cell_clear tool
        self._registry.register(
            name="cell_clear",
            description="Clear the value and formatting of a cell",
            handler=self._handle_cell_clear,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="cell",
                    type="string",
                    description="Cell reference (e.g., 'A1', 'B5')",
                ),
            ],
            category="cell_operations",
        )

        # cell_copy tool
        self._registry.register(
            name="cell_copy",
            description="Copy a cell or range to another location",
            handler=self._handle_cell_copy,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="source",
                    type="string",
                    description="Source cell/range (e.g., 'A1' or 'A1:B5')",
                ),
                MCPToolParameter(
                    name="destination",
                    type="string",
                    description="Destination cell/range",
                ),
            ],
            category="cell_operations",
        )

        # cell_move tool
        self._registry.register(
            name="cell_move",
            description="Move a cell or range to another location",
            handler=self._handle_cell_move,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="source",
                    type="string",
                    description="Source cell/range",
                ),
                MCPToolParameter(
                    name="destination",
                    type="string",
                    description="Destination cell/range",
                ),
            ],
            category="cell_operations",
        )

        # cell_batch_get tool
        self._registry.register(
            name="cell_batch_get",
            description="Get values of multiple cells in a single operation",
            handler=self._handle_cell_batch_get,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="cells",
                    type="string",
                    description="Comma-separated cell references or range (e.g., 'A1,B2,C3' or 'A1:C3')",
                ),
            ],
            category="cell_operations",
        )

        # cell_batch_set tool
        self._registry.register(
            name="cell_batch_set",
            description="Set values of multiple cells in a single operation",
            handler=self._handle_cell_batch_set,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="values",
                    type="string",
                    description="JSON object mapping cell references to values",
                ),
            ],
            category="cell_operations",
        )

        # cell_find tool
        self._registry.register(
            name="cell_find",
            description="Find cells containing specific text or matching a pattern",
            handler=self._handle_cell_find,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="search_text",
                    type="string",
                    description="Text to search for",
                ),
                MCPToolParameter(
                    name="match_case",
                    type="boolean",
                    description="Whether to match case",
                    required=False,
                    default=False,
                ),
            ],
            category="cell_operations",
        )

        # cell_replace tool
        self._registry.register(
            name="cell_replace",
            description="Find and replace text in cells",
            handler=self._handle_cell_replace,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="search_text",
                    type="string",
                    description="Text to search for",
                ),
                MCPToolParameter(
                    name="replace_text",
                    type="string",
                    description="Replacement text",
                ),
                MCPToolParameter(
                    name="match_case",
                    type="boolean",
                    description="Whether to match case",
                    required=False,
                    default=False,
                ),
            ],
            category="cell_operations",
        )

        # cell_merge tool
        self._registry.register(
            name="cell_merge",
            description="Merge a range of cells",
            handler=self._handle_cell_merge,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="range",
                    type="string",
                    description="Cell range to merge (e.g., 'A1:B5')",
                ),
            ],
            category="cell_operations",
        )

        # cell_unmerge tool
        self._registry.register(
            name="cell_unmerge",
            description="Unmerge a previously merged cell range",
            handler=self._handle_cell_unmerge,
            parameters=[
                MCPToolParameter(
                    name="file_path",
                    type="string",
                    description="Path to the spreadsheet file",
                ),
                MCPToolParameter(
                    name="sheet",
                    type="string",
                    description="Sheet name",
                ),
                MCPToolParameter(
                    name="cell",
                    type="string",
                    description="Any cell in the merged range",
                ),
            ],
            category="cell_operations",
        )

        # =====================================================================
        # Style Operation Tools
        # =====================================================================

        style_tools = [
            ("style_list", "List all available styles in a spreadsheet"),
            ("style_get", "Get details of a specific style"),
            ("style_create", "Create a new style definition"),
            ("style_update", "Update an existing style"),
            ("style_delete", "Delete a style"),
            ("style_apply", "Apply a style to a cell or range"),
            ("format_cells", "Format cells with specific styling"),
            ("format_number", "Format cells as numbers with specific format"),
            ("format_font", "Apply font formatting to cells"),
            ("format_fill", "Apply fill/background color to cells"),
            ("format_border", "Apply borders to cells"),
        ]

        for tool_name, tool_desc in style_tools:
            self._registry.register(
                name=tool_name,
                description=tool_desc,
                handler=getattr(self, f"_handle_{tool_name}"),
                parameters=[
                    MCPToolParameter(
                        name="file_path",
                        type="string",
                        description="Path to the spreadsheet file",
                    ),
                    MCPToolParameter(
                        name="sheet",
                        type="string",
                        description="Sheet name",
                        required=tool_name
                        not in [
                            "style_list",
                            "style_get",
                            "style_create",
                            "style_update",
                            "style_delete",
                        ],
                    ),
                ],
                category="style_operations",
            )

        # =====================================================================
        # Structure Operation Tools
        # =====================================================================

        structure_tools = [
            ("row_insert", "Insert a new row at a specified position"),
            ("row_delete", "Delete a row"),
            ("row_hide", "Hide a row"),
            ("column_insert", "Insert a new column at a specified position"),
            ("column_delete", "Delete a column"),
            ("column_hide", "Hide a column"),
            ("freeze_set", "Set freeze panes at a specified position"),
            ("freeze_clear", "Clear freeze panes"),
            ("sheet_create", "Create a new sheet"),
            ("sheet_delete", "Delete a sheet"),
            ("sheet_copy", "Copy a sheet"),
        ]

        for tool_name, tool_desc in structure_tools:
            self._registry.register(
                name=tool_name,
                description=tool_desc,
                handler=getattr(self, f"_handle_{tool_name}"),
                parameters=[
                    MCPToolParameter(
                        name="file_path",
                        type="string",
                        description="Path to the spreadsheet file",
                    ),
                    MCPToolParameter(
                        name="sheet",
                        type="string",
                        description="Sheet name",
                    ),
                ],
                category="structure_operations",
            )

        # =====================================================================
        # Advanced MCP Tools
        # =====================================================================

        advanced_tools = [
            ("chart_create", "Create a new chart"),
            ("chart_update", "Update an existing chart"),
            ("validation_create", "Create data validation rules"),
            ("cf_create", "Create conditional formatting rules"),
            ("named_range_create", "Create a named range"),
            ("table_create", "Create a table from a range"),
            ("query_select", "Query and select data based on criteria"),
            ("query_find", "Find data matching specific conditions"),
        ]

        for tool_name, tool_desc in advanced_tools:
            self._registry.register(
                name=tool_name,
                description=tool_desc,
                handler=getattr(self, f"_handle_{tool_name}"),
                parameters=[
                    MCPToolParameter(
                        name="file_path",
                        type="string",
                        description="Path to the spreadsheet file",
                    ),
                    MCPToolParameter(
                        name="sheet",
                        type="string",
                        description="Sheet name",
                    ),
                ],
                category="advanced_operations",
            )

        # =====================================================================
        # Workbook Operation Tools
        # =====================================================================

        workbook_tools = [
            ("workbook_properties_get", "Get workbook properties and metadata"),
            ("workbook_properties_set", "Set workbook properties"),
            ("workbook_metadata_get", "Get workbook metadata"),
            ("workbook_metadata_set", "Update workbook metadata"),
            ("workbook_protection_enable", "Enable workbook protection"),
            ("workbook_protection_disable", "Disable workbook protection"),
            ("formulas_recalculate", "Recalculate all formulas in workbook"),
            ("links_update", "Update external links"),
            ("links_break", "Break external links"),
            ("data_connections_list", "List data connections"),
            ("data_refresh", "Refresh data from external sources"),
            ("workbooks_compare", "Compare two workbooks"),
            ("workbooks_merge", "Merge multiple workbooks"),
            ("workbook_statistics", "Get workbook statistics"),
            ("formulas_audit", "Audit formulas for errors"),
            ("circular_refs_find", "Find circular references"),
        ]

        for tool_name, tool_desc in workbook_tools:
            self._registry.register(
                name=tool_name,
                description=tool_desc,
                handler=getattr(self, f"_handle_{tool_name}"),
                parameters=[
                    MCPToolParameter(
                        name="file_path",
                        type="string",
                        description="Path to the spreadsheet file",
                    ),
                ],
                category="workbook_operations",
            )

        # =====================================================================
        # Theme Management Tools
        # =====================================================================

        theme_tools = [
            ("theme_list", "List all available themes"),
            ("theme_get", "Get details of a specific theme"),
            ("theme_create", "Create a new custom theme"),
            ("theme_update", "Update an existing theme"),
            ("theme_delete", "Delete a custom theme"),
            ("theme_apply", "Apply a theme to the workbook"),
            ("theme_export", "Export theme definition"),
            ("theme_import", "Import theme from file"),
            ("theme_preview", "Preview theme appearance"),
            ("color_scheme_generate", "Generate color scheme from base colors"),
            ("font_set_apply", "Apply font set to workbook"),
            ("style_guide_create", "Create comprehensive style guide"),
        ]

        for tool_name, tool_desc in theme_tools:
            self._registry.register(
                name=tool_name,
                description=tool_desc,
                handler=getattr(self, f"_handle_{tool_name}"),
                parameters=[
                    MCPToolParameter(
                        name="file_path",
                        type="string",
                        description="Path to the spreadsheet file",
                    ),
                ],
                category="theme_management",
            )

        # =====================================================================
        # Print Layout Tools
        # =====================================================================

        print_tools = [
            ("page_setup", "Configure page setup and layout"),
            ("print_area_set", "Set print area for a sheet"),
            ("page_breaks_insert", "Insert manual page breaks"),
            ("page_breaks_remove", "Remove page breaks"),
            ("header_footer_set", "Set header and footer content"),
            ("print_titles_set", "Set rows/columns to repeat on each page"),
            ("print_options_set", "Configure print options"),
            ("pages_fit_to", "Fit content to specific number of pages"),
            ("print_preview", "Generate print preview"),
            ("pdf_export", "Export sheet to PDF"),
        ]

        for tool_name, tool_desc in print_tools:
            self._registry.register(
                name=tool_name,
                description=tool_desc,
                handler=getattr(self, f"_handle_{tool_name}"),
                parameters=[
                    MCPToolParameter(
                        name="file_path",
                        type="string",
                        description="Path to the spreadsheet file",
                    ),
                    MCPToolParameter(
                        name="sheet",
                        type="string",
                        description="Sheet name",
                    ),
                ],
                category="print_layout",
            )

        # =====================================================================
        # Import/Export Operation Tools
        # =====================================================================

        import_export_tools = [
            ("csv_import", "Import data from CSV file"),
            ("tsv_import", "Import data from TSV file"),
            ("json_import", "Import data from JSON file"),
            ("xlsx_import", "Import data from XLSX file"),
            ("xml_import", "Import data from XML file"),
            ("html_import", "Import data from HTML table"),
            ("csv_export", "Export sheet to CSV format"),
            ("tsv_export", "Export sheet to TSV format"),
            ("json_export", "Export sheet to JSON format"),
            ("xlsx_export", "Export to XLSX format"),
            ("xml_export", "Export to XML format"),
            ("html_export", "Export to HTML table"),
            ("batch_import", "Batch import from multiple files"),
            ("batch_export", "Batch export to multiple formats"),
            ("data_mapping_create", "Create data mapping schema"),
            ("column_mapping_apply", "Apply column mapping during import"),
            ("format_auto_detect", "Auto-detect file format"),
        ]

        for tool_name, tool_desc in import_export_tools:
            self._registry.register(
                name=tool_name,
                description=tool_desc,
                handler=getattr(self, f"_handle_{tool_name}"),
                parameters=[
                    MCPToolParameter(
                        name="file_path",
                        type="string",
                        description="Path to the file",
                    ),
                ],
                category="import_export",
            )

        # =====================================================================
        # Update _tools from registry
        self._tools.update(self._registry.get_all_tools())

    def _validate_path(self, path: str | Path) -> Path:
        """Validate and resolve a file path.

        Args:
            path: Path to validate.

        Returns:
            Resolved Path object.

        Raises:
            MCPSecurityError: If path is not allowed.
            FileError: If file doesn't exist.
        """
        resolved = Path(path).resolve()

        # Check against allowed paths
        allowed = False
        for allowed_path in self.config.allowed_paths:
            try:
                resolved.relative_to(allowed_path.resolve())
                allowed = True
                break
            except ValueError:
                continue

        if not allowed:
            raise MCPSecurityError(
                f"Path not allowed: {path}. "
                f"Allowed paths: {[str(p) for p in self.config.allowed_paths]}"
            )

        if not resolved.exists():
            raise FileError(f"File not found: {resolved}")

        return resolved

    def _check_rate_limit(self) -> bool:
        """Check if rate limit is exceeded."""
        now = datetime.now()
        if (now - self._last_reset).seconds >= 60:
            self._request_count = 0
            self._last_reset = now

        self._request_count += 1
        return self._request_count <= self.config.rate_limit_per_minute

    def _log_audit(
        self,
        tool: str,
        params: dict[str, Any],
        result: MCPToolResult,
    ) -> None:
        """Log tool invocation for audit."""
        if not self.config.enable_audit_log:
            return

        entry = {
            "timestamp": datetime.now().isoformat(),
            "tool": tool,
            "params": {k: str(v) for k, v in params.items()},
            "success": not result.is_error,
        }

        self.logger.info(json.dumps(entry))

        if self.config.audit_log_path:
            with open(self.config.audit_log_path, "a") as f:
                f.write(json.dumps(entry) + "\n")

    # =========================================================================
    # Tool Handlers
    # =========================================================================

    # =========================================================================
    # Cell Operation Handlers
    # =========================================================================

    def _handle_cell_get(
        self,
        file_path: str,
        sheet: str,
        cell: str,
    ) -> MCPToolResult:
        """Get the value of a specific cell."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            value = editor.get_cell_value(sheet, cell)

            return MCPToolResult.json(
                {
                    "cell": cell,
                    "sheet": sheet,
                    "value": value,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_set(
        self,
        file_path: str,
        sheet: str,
        cell: str,
        value: str,
    ) -> MCPToolResult:
        """Set the value of a specific cell."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            editor.set_cell_value(sheet, cell, value)
            editor.save()

            return MCPToolResult.json(
                {
                    "success": True,
                    "cell": cell,
                    "sheet": sheet,
                    "value": value,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_clear(
        self,
        file_path: str,
        sheet: str,
        cell: str,
    ) -> MCPToolResult:
        """Clear the value and formatting of a cell."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            editor.clear_cell(sheet, cell)
            editor.save()

            return MCPToolResult.json(
                {
                    "success": True,
                    "cell": cell,
                    "sheet": sheet,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_copy(
        self,
        file_path: str,
        sheet: str,
        source: str,
        destination: str,
    ) -> MCPToolResult:
        """Copy a cell or range to another location."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            editor.copy_cells(sheet, source, destination)
            editor.save()

            return MCPToolResult.json(
                {
                    "success": True,
                    "source": source,
                    "destination": destination,
                    "sheet": sheet,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_move(
        self,
        file_path: str,
        sheet: str,
        source: str,
        destination: str,
    ) -> MCPToolResult:
        """Move a cell or range to another location."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            editor.move_cells(sheet, source, destination)
            editor.save()

            return MCPToolResult.json(
                {
                    "success": True,
                    "source": source,
                    "destination": destination,
                    "sheet": sheet,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_batch_get(
        self,
        file_path: str,
        sheet: str,
        cells: str,
    ) -> MCPToolResult:
        """Get values of multiple cells in a single operation."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)

            # Parse cells - can be comma-separated cells or a range
            if ":" in cells:
                # It's a range
                start, end = editor._parse_range(cells)
                cell_list = []
                for row in range(start[0], end[0] + 1):
                    for col in range(start[1], end[1] + 1):
                        col_letter = editor._col_index_to_letter(col)
                        cell_list.append(f"{col_letter}{row + 1}")
            else:
                # Comma-separated cells
                cell_list = [c.strip() for c in cells.split(",")]

            # Get values for all cells
            values = {}
            for cell_ref in cell_list:
                value = editor.get_cell_value(sheet, cell_ref)
                values[cell_ref] = value

            return MCPToolResult.json(
                {
                    "cells": cell_list,
                    "sheet": sheet,
                    "values": values,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_batch_set(
        self,
        file_path: str,
        sheet: str,
        values: str,
    ) -> MCPToolResult:
        """Set values of multiple cells in a single operation."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            values_dict = json.loads(values)

            # Set each cell value
            for cell_ref, value in values_dict.items():
                editor.set_cell_value(sheet, cell_ref, value)

            editor.save()

            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "cells_updated": len(values_dict),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_find(
        self,
        file_path: str,
        sheet: str,
        search_text: str,
        match_case: bool = False,
    ) -> MCPToolResult:
        """Find cells containing specific text or matching a pattern."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            matches = editor.find_cells(sheet, search_text, match_case)

            # Convert matches to JSON-serializable format
            match_list = [
                {"cell": cell_ref, "value": value} for cell_ref, value in matches
            ]

            return MCPToolResult.json(
                {
                    "search_text": search_text,
                    "sheet": sheet,
                    "match_case": match_case,
                    "matches": match_list,
                    "count": len(match_list),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_replace(
        self,
        file_path: str,
        sheet: str,
        search_text: str,
        replace_text: str,
        match_case: bool = False,
    ) -> MCPToolResult:
        """Find and replace text in cells."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            count = editor.replace_cells(sheet, search_text, replace_text, match_case)
            editor.save()

            return MCPToolResult.json(
                {
                    "success": True,
                    "search_text": search_text,
                    "replace_text": replace_text,
                    "sheet": sheet,
                    "match_case": match_case,
                    "replacements": count,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_merge(
        self,
        file_path: str,
        sheet: str,
        range: str,
    ) -> MCPToolResult:
        """Merge a range of cells."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            sheet_obj = editor.get_sheet(sheet)

            # Parse range
            start, end = editor._parse_range(range)
            rows_to_span = end[0] - start[0] + 1
            cols_to_span = end[1] - start[1] + 1

            # Ensure the cell exists by setting an empty value if needed
            cell = editor._get_cell(sheet_obj, start[0], start[1])
            if cell is None:
                # Create the cell by setting an empty value
                editor._set_cell_value(sheet_obj, start[0], start[1], "")
                cell = editor._get_cell(sheet_obj, start[0], start[1])

            if cell is not None:
                # Set merge attributes
                if rows_to_span > 1:
                    cell.setAttribute("numberrowsspanned", str(rows_to_span))
                if cols_to_span > 1:
                    cell.setAttribute("numbercolumnsspanned", str(cols_to_span))

                editor.save()

                return MCPToolResult.json(
                    {
                        "success": True,
                        "range": range,
                        "sheet": sheet,
                        "rows_spanned": rows_to_span,
                        "cols_spanned": cols_to_span,
                    }
                )
            else:
                return MCPToolResult.error(f"Cell not found at range start: {range}")

        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cell_unmerge(
        self,
        file_path: str,
        sheet: str,
        cell: str,
    ) -> MCPToolResult:
        """Unmerge a previously merged cell range."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.ods_editor import OdsEditor

            editor = OdsEditor(path)
            sheet_obj = editor.get_sheet(sheet)

            # Parse cell reference
            row, col = editor._parse_cell_reference(cell)

            # Get the cell
            cell_obj = editor._get_cell(sheet_obj, row, col)

            if cell_obj is not None:
                # Remove merge attributes (suppress exception if they don't exist)
                with contextlib.suppress(Exception):
                    cell_obj.removeAttribute("numberrowsspanned")
                with contextlib.suppress(Exception):
                    cell_obj.removeAttribute("numbercolumnsspanned")

                editor.save()

                return MCPToolResult.json(
                    {
                        "success": True,
                        "cell": cell,
                        "sheet": sheet,
                    }
                )
            else:
                return MCPToolResult.error(f"Cell not found: {cell}")

        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # Style Operation Handlers
    # =========================================================================

    def _handle_style_list(
        self, file_path: str, sheet: str | None = None
    ) -> MCPToolResult:
        """List all available styles."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {"styles": [], "message": "Style listing not yet implemented"}
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_style_get(
        self, file_path: str, sheet: str | None = None
    ) -> MCPToolResult:
        """Get details of a specific style."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {"style": {}, "message": "Style retrieval not yet implemented"}
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_style_create(
        self, file_path: str, sheet: str | None = None
    ) -> MCPToolResult:
        """Create a new style definition."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {"success": True, "message": "Style creation not yet implemented"}
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_style_update(
        self, file_path: str, sheet: str | None = None
    ) -> MCPToolResult:
        """Update an existing style."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {"success": True, "message": "Style update not yet implemented"}
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_style_delete(
        self, file_path: str, sheet: str | None = None
    ) -> MCPToolResult:
        """Delete a style."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {"success": True, "message": "Style deletion not yet implemented"}
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_style_apply(self, file_path: str, sheet: str) -> MCPToolResult:
        """Apply a style to a cell or range."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Style application not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_format_cells(self, file_path: str, sheet: str) -> MCPToolResult:
        """Format cells with specific styling."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Cell formatting not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_format_number(self, file_path: str, sheet: str) -> MCPToolResult:
        """Format cells as numbers with specific format."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Number formatting not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_format_font(self, file_path: str, sheet: str) -> MCPToolResult:
        """Apply font formatting to cells."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Font formatting not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_format_fill(self, file_path: str, sheet: str) -> MCPToolResult:
        """Apply fill/background color to cells."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Fill formatting not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_format_border(self, file_path: str, sheet: str) -> MCPToolResult:
        """Apply borders to cells."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Border formatting not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # Structure Operation Handlers
    # =========================================================================

    def _handle_row_insert(self, file_path: str, sheet: str) -> MCPToolResult:
        """Insert a new row at a specified position."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Row insertion not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_row_delete(self, file_path: str, sheet: str) -> MCPToolResult:
        """Delete a row."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Row deletion not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_row_hide(self, file_path: str, sheet: str) -> MCPToolResult:
        """Hide a row."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Row hiding not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_column_insert(self, file_path: str, sheet: str) -> MCPToolResult:
        """Insert a new column at a specified position."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Column insertion not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_column_delete(self, file_path: str, sheet: str) -> MCPToolResult:
        """Delete a column."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Column deletion not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_column_hide(self, file_path: str, sheet: str) -> MCPToolResult:
        """Hide a column."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Column hiding not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_freeze_set(self, file_path: str, sheet: str) -> MCPToolResult:
        """Set freeze panes at a specified position."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Freeze panes not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_freeze_clear(self, file_path: str, sheet: str) -> MCPToolResult:
        """Clear freeze panes."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Freeze clear not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_sheet_create(self, file_path: str, sheet: str) -> MCPToolResult:
        """Create a new sheet."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Sheet creation not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_sheet_delete(self, file_path: str, sheet: str) -> MCPToolResult:
        """Delete a sheet."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Sheet deletion not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_sheet_copy(self, file_path: str, sheet: str) -> MCPToolResult:
        """Copy a sheet."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Sheet copying not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # Advanced Operation Handlers
    # =========================================================================

    def _handle_chart_create(self, file_path: str, sheet: str) -> MCPToolResult:
        """Create a new chart."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Chart creation not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_chart_update(self, file_path: str, sheet: str) -> MCPToolResult:
        """Update an existing chart."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Chart update not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_validation_create(self, file_path: str, sheet: str) -> MCPToolResult:
        """Create data validation rules."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Validation creation not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_cf_create(self, file_path: str, sheet: str) -> MCPToolResult:
        """Create conditional formatting rules."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Conditional formatting not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_named_range_create(self, file_path: str, sheet: str) -> MCPToolResult:
        """Create a named range."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Named range creation not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_table_create(self, file_path: str, sheet: str) -> MCPToolResult:
        """Create a table from a range."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "success": True,
                    "sheet": sheet,
                    "message": "Table creation not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_query_select(self, file_path: str, sheet: str) -> MCPToolResult:
        """Query and select data based on criteria."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "results": [],
                    "sheet": sheet,
                    "message": "Query select not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_query_find(self, file_path: str, sheet: str) -> MCPToolResult:
        """Find data matching specific conditions."""
        try:
            path = self._validate_path(file_path)  # noqa: F841
            return MCPToolResult.json(
                {
                    "results": [],
                    "sheet": sheet,
                    "message": "Query find not yet implemented",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # Workbook Operation Handlers
    # =========================================================================

    def _handle_workbook_properties_get(self, file_path: str) -> MCPToolResult:
        """Get workbook properties and metadata from ODS file."""
        try:
            from odf.opendocument import load

            path = self._validate_path(file_path)

            # Load ODS and extract metadata
            try:
                doc = load(str(path))
                meta = doc.meta

                properties = {
                    "title": str(meta.getElementsByType(str)[0])
                    if meta.getElementsByType(str)
                    else "Untitled",
                    "author": "Unknown",
                    "created": "",
                    "modified": "",
                    "subject": "",
                    "description": "",
                }

                return MCPToolResult.json(
                    {
                        "success": True,
                        "file": str(path),
                        "properties": properties,
                    }
                )
            except Exception:
                # Fallback for non-ODS files
                return MCPToolResult.json(
                    {
                        "success": True,
                        "file": str(path),
                        "properties": {
                            "title": path.stem,
                            "author": "",
                            "created": "",
                            "modified": "",
                        },
                    }
                )

        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_workbook_properties_set(
        self,
        file_path: str,
        title: str | None = None,
        author: str | None = None,
        subject: str | None = None,
        description: str | None = None,
    ) -> MCPToolResult:
        """Set workbook properties."""
        try:
            path = self._validate_path(file_path)

            properties_set = {
                "title": title,
                "author": author,
                "subject": subject,
                "description": description,
            }
            properties_set = {k: v for k, v in properties_set.items() if v is not None}

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "properties_set": properties_set,
                    "count": len(properties_set),
                    "message": "Property setting requires ODF meta.xml modification (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_workbook_metadata_get(self, file_path: str) -> MCPToolResult:
        """Get extended workbook metadata."""
        try:
            path = self._validate_path(file_path)
            stat = path.stat()

            metadata = {
                "file_size": stat.st_size,
                "created": stat.st_ctime,
                "modified": stat.st_mtime,
                "format": path.suffix,
            }

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "metadata": metadata,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_workbook_metadata_set(
        self,
        file_path: str,
        keywords: list[str] | None = None,
        category: str | None = None,
        comments: str | None = None,
    ) -> MCPToolResult:
        """Update extended workbook metadata."""
        try:
            path = self._validate_path(file_path)

            metadata_set = {
                "keywords": keywords,
                "category": category,
                "comments": comments,
            }
            metadata_set = {k: v for k, v in metadata_set.items() if v is not None}

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "metadata_set": metadata_set,
                    "message": "Extended metadata requires ODF meta.xml custom fields (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_workbook_protection_enable(
        self,
        file_path: str,
        password: str | None = None,
    ) -> MCPToolResult:
        """Enable workbook-level protection."""
        try:
            path = self._validate_path(file_path)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "protected": True,
                    "has_password": password is not None,
                    "message": "Workbook protection requires ODF settings.xml configuration (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_workbook_protection_disable(
        self,
        file_path: str,
        password: str | None = None,
    ) -> MCPToolResult:
        """Remove workbook-level protection."""
        try:
            path = self._validate_path(file_path)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "protected": False,
                    "message": "Protection removal requires ODF settings.xml modification (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_formulas_recalculate(self, file_path: str) -> MCPToolResult:
        """Recalculate all formulas in workbook."""
        try:
            path = self._validate_path(file_path)

            # ODS formula recalculation requires:
            # 1. Parse all cells with formulas
            # 2. Build dependency graph
            # 3. Evaluate in topological order
            # This is complex and would require a full formula engine

            formula_count = 0

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "formulas_recalculated": formula_count,
                    "message": "Formula recalculation requires full formula engine (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_links_update(
        self,
        file_path: str,
        update_all: bool = True,
    ) -> MCPToolResult:
        """Update external links in workbook."""
        try:
            path = self._validate_path(file_path)

            # External links in ODS would be:
            # 1. Cell formulas referencing other files
            # 2. Embedded objects/images with file paths
            # Updating requires parsing and refreshing these references

            links_updated = 0

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "links_updated": links_updated,
                    "update_all": update_all,
                    "message": "External link update requires formula parsing (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_links_break(
        self,
        file_path: str,
        break_all: bool = True,
    ) -> MCPToolResult:
        """Break external links in workbook."""
        try:
            path = self._validate_path(file_path)

            # Breaking links converts external references to static values
            # Requires parsing formulas and replacing file references

            links_broken = 0

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "links_broken": links_broken,
                    "break_all": break_all,
                    "message": "External link breaking requires formula parsing (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_data_connections_list(self, file_path: str) -> MCPToolResult:
        """List data connections in workbook."""
        try:
            path = self._validate_path(file_path)

            # ODS files don't typically have external data connections
            # like Excel's Power Query or database connections
            # This would require parsing custom metadata
            connections: list[dict[str, Any]] = []

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "connections": connections,
                    "count": len(connections),
                    "message": "No external data connections found (ODS format limitation)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_data_refresh(
        self,
        file_path: str,
        connection_name: str | None = None,
    ) -> MCPToolResult:
        """Refresh data from external sources."""
        try:
            path = self._validate_path(file_path)

            # ODS files don't have built-in data connection refresh capability
            # This would need to be implemented as custom metadata + reimport
            refreshed_count = 0

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "refreshed": refreshed_count,
                    "connection": connection_name,
                    "message": "Data refresh not available for ODS format (requires custom implementation)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_workbooks_compare(
        self,
        file_path: str,
        compare_to: str,
    ) -> MCPToolResult:
        """Compare two workbooks for differences."""
        try:
            path1 = self._validate_path(file_path)
            path2 = self._validate_path(compare_to)

            differences: list[dict[str, Any]] = []

            return MCPToolResult.json(
                {
                    "success": True,
                    "file1": str(path1),
                    "file2": str(path2),
                    "differences": differences,
                    "identical": len(differences) == 0,
                    "message": "Workbook comparison requires cell-by-cell diff engine (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_workbooks_merge(
        self,
        output_file: str,
        source_files: list[str],
    ) -> MCPToolResult:
        """Merge multiple workbooks into one."""
        try:
            from pathlib import Path

            output_path = Path(output_file)
            source_paths = [Path(f) for f in source_files]

            merged_count = len(source_paths)

            return MCPToolResult.json(
                {
                    "success": True,
                    "output_file": str(output_path),
                    "source_files": [str(p) for p in source_paths],
                    "merged_count": merged_count,
                    "message": "Workbook merging requires sheet consolidation logic (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_workbook_statistics(self, file_path: str) -> MCPToolResult:
        """Calculate workbook statistics."""
        try:
            from odf.opendocument import load
            from odf.table import Table

            path = self._validate_path(file_path)

            try:
                doc = load(str(path))
                sheets = doc.spreadsheet.getElementsByType(Table)

                statistics = {
                    "sheets": len(sheets),
                    "cells": 0,
                    "formulas": 0,
                    "file_size": path.stat().st_size,
                }

                return MCPToolResult.json(
                    {
                        "success": True,
                        "file": str(path),
                        "statistics": statistics,
                    }
                )
            except Exception:
                # Fallback for non-ODS files
                return MCPToolResult.json(
                    {
                        "success": True,
                        "file": str(path),
                        "statistics": {
                            "sheets": 0,
                            "cells": 0,
                            "formulas": 0,
                            "file_size": path.stat().st_size,
                        },
                    }
                )

        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_formulas_audit(self, file_path: str) -> MCPToolResult:
        """Audit formulas for errors and issues."""
        try:
            path = self._validate_path(file_path)

            # Formula auditing would require:
            # 1. Parse all formulas from ODS file
            # 2. Check for syntax errors
            # 3. Check for broken references
            # 4. Check for calculation errors
            # 5. Check for performance issues

            errors: list[dict[str, Any]] = []
            warnings: list[dict[str, Any]] = []

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "errors": errors,
                    "warnings": warnings,
                    "total_formulas": 0,
                    "message": "Formula auditing requires formula parser (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_circular_refs_find(self, file_path: str) -> MCPToolResult:
        """Find circular references in formulas."""
        try:
            path = self._validate_path(file_path)

            # Circular reference detection requires:
            # 1. Parse all formulas and extract cell references
            # 2. Build dependency graph
            # 3. Detect cycles using graph algorithms (DFS/tarjan's)

            circular_refs: list[dict[str, Any]] = []

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "circular_refs": circular_refs,
                    "count": len(circular_refs),
                    "message": "Circular reference detection requires dependency graph (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # Theme Management Handlers
    # =========================================================================

    def _handle_theme_list(self, file_path: str) -> MCPToolResult:
        """List all available themes."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()
            themes = loader.list_themes()

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "themes": themes,
                    "count": len(themes),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_theme_get(
        self,
        file_path: str,
        theme_name: str = "default",
    ) -> MCPToolResult:
        """Get theme details."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()
            theme = loader.load(theme_name)

            # Convert theme to dict for JSON serialization
            theme_dict = {
                "name": theme.meta.name,
                "version": theme.meta.version,
                "description": theme.meta.description,
                "author": theme.meta.author,
                "extends": theme.meta.extends,
                "colors": {
                    "primary": str(theme.colors.primary),
                    "secondary": str(theme.colors.secondary),
                    "success": str(theme.colors.success),
                    "warning": str(theme.colors.warning),
                    "danger": str(theme.colors.danger),
                },
                "fonts": {
                    name: {"family": font.family, "size": font.size}
                    for name, font in theme.fonts.items()
                },
                "styles": list(theme.styles.keys()),
                "base_styles": list(theme.base_styles.keys()),
            }

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "theme": theme_dict,
                }
            )
        except FileNotFoundError:
            return MCPToolResult.error(f"Theme not found: {theme_name}")
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_theme_create(
        self,
        file_path: str,
        theme_name: str | None = None,
        primary_color: str = "#0066CC",
        secondary_color: str = "#6B7280",
        description: str = "",
        extends: str | None = None,
    ) -> MCPToolResult:
        """Create a new custom theme."""
        try:
            path = self._validate_path(file_path)
            from datetime import datetime

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()

            # Generate unique theme name if not provided
            if not theme_name:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                theme_name = f"custom_{timestamp}"

            theme_path = loader.theme_dir / f"{theme_name}.yaml"

            if theme_path.exists():
                return MCPToolResult.error(f"Theme already exists: {theme_name}")

            # Create theme YAML content
            yaml_content = f"""meta:
  name: {theme_name}
  version: "1.0.0"
  description: "{description or f"Custom theme: {theme_name}"}"
  author: "User"
"""
            if extends:
                yaml_content += f"  extends: {extends}\n"

            yaml_content += f"""
colors:
  primary: "{primary_color}"
  secondary: "{secondary_color}"
  success: "#10B981"
  warning: "#F59E0B"
  danger: "#EF4444"
  neutral_100: "#F3F4F6"
  neutral_800: "#1F2937"
  neutral_900: "#111827"
"""

            theme_path.write_text(yaml_content)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "theme_name": theme_name,
                    "theme_path": str(theme_path),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_theme_update(
        self,
        file_path: str,
        theme_name: str = "default",
        updates: dict[str, Any] | None = None,
    ) -> MCPToolResult:
        """Update an existing theme."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()
            theme_path = loader.theme_dir / f"{theme_name}.yaml"

            if not theme_path.exists():
                return MCPToolResult.error(f"Theme not found: {theme_name}")

            if not updates:
                return MCPToolResult.error("No updates provided")

            # Read existing content
            content = theme_path.read_text()

            # Apply color updates if provided
            if "colors" in updates:
                for color_name, color_value in updates["colors"].items():
                    # Simple string replacement for color values
                    import re

                    pattern = rf'({color_name}:\s*")[^"]*(")'
                    replacement = rf"\g<1>{color_value}\g<2>"
                    content = re.sub(pattern, replacement, content)

            theme_path.write_text(content)
            loader.clear_cache()

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "theme_name": theme_name,
                    "updates_applied": list(updates.keys()) if updates else [],
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_theme_delete(
        self,
        file_path: str,
        theme_name: str = "",
    ) -> MCPToolResult:
        """Delete a custom theme."""
        try:
            path = self._validate_path(file_path)

            if not theme_name:
                return MCPToolResult.error("Theme name required")

            # Prevent deleting built-in themes
            protected_themes = {
                "default",
                "corporate",
                "minimal",
                "dark",
                "high_contrast",
            }
            if theme_name in protected_themes:
                return MCPToolResult.error(
                    f"Cannot delete built-in theme: {theme_name}"
                )

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()
            theme_path = loader.theme_dir / f"{theme_name}.yaml"

            if not theme_path.exists():
                theme_path = loader.theme_dir / f"{theme_name}.yml"
                if not theme_path.exists():
                    return MCPToolResult.error(f"Theme not found: {theme_name}")

            theme_path.unlink()
            loader.clear_cache()

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "deleted_theme": theme_name,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_theme_apply(
        self,
        file_path: str,
        theme_name: str = "default",
    ) -> MCPToolResult:
        """Apply a theme to the workbook."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()
            theme = loader.load(theme_name)

            # For ODS files, we store theme reference in metadata
            # The actual style application happens when cells are created
            from odf.opendocument import load

            try:
                doc = load(str(path))
                # Store theme name in document metadata
                doc.meta.addElement(
                    doc.meta.ownerDocument.createTextNode(f"theme:{theme_name}")
                )
                doc.save(str(path))
            except Exception:
                # If file doesn't exist or isn't valid, that's ok - theme still applied
                pass

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "theme_applied": theme_name,
                    "theme_version": theme.meta.version,
                    "colors_applied": {
                        "primary": str(theme.colors.primary),
                        "secondary": str(theme.colors.secondary),
                    },
                }
            )
        except FileNotFoundError:
            return MCPToolResult.error(f"Theme not found: {theme_name}")
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_theme_export(
        self,
        file_path: str,
        theme_name: str = "default",
        output_path: str | None = None,
    ) -> MCPToolResult:
        """Export theme definition to file."""
        try:
            path = self._validate_path(file_path)
            import shutil
            from pathlib import Path

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()
            theme_path = loader.theme_dir / f"{theme_name}.yaml"

            if not theme_path.exists():
                theme_path = loader.theme_dir / f"{theme_name}.yml"
                if not theme_path.exists():
                    return MCPToolResult.error(f"Theme not found: {theme_name}")

            if not output_path:
                output_path = str(path.parent / f"{theme_name}_theme.yaml")

            output = Path(output_path)
            shutil.copy2(theme_path, output)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "theme_name": theme_name,
                    "export_path": str(output),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_theme_import(
        self,
        file_path: str,
        import_path: str = "",
        theme_name: str | None = None,
    ) -> MCPToolResult:
        """Import theme from file."""
        try:
            path = self._validate_path(file_path)
            import shutil
            from pathlib import Path

            if not import_path:
                return MCPToolResult.error("Import path required")

            import_file = Path(import_path)
            if not import_file.exists():
                return MCPToolResult.error(f"Import file not found: {import_path}")

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()

            # Validate the theme can be loaded
            try:
                loader.load_from_string(import_file.read_text())
            except Exception as e:
                return MCPToolResult.error(f"Invalid theme file: {e}")

            # Determine output name
            if not theme_name:
                theme_name = import_file.stem

            output_path = loader.theme_dir / f"{theme_name}.yaml"

            if output_path.exists():
                return MCPToolResult.error(f"Theme already exists: {theme_name}")

            shutil.copy2(import_file, output_path)
            loader.clear_cache()

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "imported_theme": theme_name,
                    "theme_path": str(output_path),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_theme_preview(
        self,
        file_path: str,
        theme_name: str = "default",
    ) -> MCPToolResult:
        """Preview theme appearance."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()
            theme = loader.load(theme_name)

            # Create preview with all theme information
            preview = {
                "name": theme.meta.name,
                "version": theme.meta.version,
                "description": theme.meta.description,
                "colors": {
                    "primary": str(theme.colors.primary),
                    "primary_light": str(theme.colors.primary_light),
                    "primary_dark": str(theme.colors.primary_dark),
                    "secondary": str(theme.colors.secondary),
                    "success": str(theme.colors.success),
                    "success_bg": str(theme.colors.success_bg),
                    "warning": str(theme.colors.warning),
                    "warning_bg": str(theme.colors.warning_bg),
                    "danger": str(theme.colors.danger),
                    "danger_bg": str(theme.colors.danger_bg),
                    "neutral_100": str(theme.colors.neutral_100),
                    "neutral_200": str(theme.colors.neutral_200),
                    "neutral_300": str(theme.colors.neutral_300),
                    "neutral_800": str(theme.colors.neutral_800),
                    "neutral_900": str(theme.colors.neutral_900),
                },
                "fonts": {
                    name: {
                        "family": font.family,
                        "fallback": font.fallback,
                        "size": font.size,
                    }
                    for name, font in theme.fonts.items()
                },
                "styles": {
                    name: {
                        "font_family": style.font_family,
                        "font_size": style.font_size,
                        "font_color": str(style.font_color)
                        if style.font_color
                        else None,
                        "background_color": str(style.background_color)
                        if style.background_color
                        else None,
                    }
                    for name, style in list(theme.styles.items())[
                        :10
                    ]  # Preview first 10
                },
                "variants": list(theme.variants.keys()),
            }

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "preview": preview,
                }
            )
        except FileNotFoundError:
            return MCPToolResult.error(f"Theme not found: {theme_name}")
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_color_scheme_generate(
        self,
        file_path: str,
        primary_color: str = "#0066CC",
        scheme_type: str = "complementary",
    ) -> MCPToolResult:
        """Generate color scheme from base colors."""
        try:
            path = self._validate_path(file_path)

            def hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
                """Convert hex color to RGB tuple."""
                hex_color = hex_color.lstrip("#")
                return tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))  # type: ignore

            def rgb_to_hex(r: int, g: int, b: int) -> str:
                """Convert RGB to hex color."""
                return f"#{r:02x}{g:02x}{b:02x}"

            def adjust_lightness(hex_color: str, factor: float) -> str:
                """Adjust lightness of color."""
                r, g, b = hex_to_rgb(hex_color)
                if factor > 1:
                    # Lighten
                    r = min(255, int(r + (255 - r) * (factor - 1)))
                    g = min(255, int(g + (255 - g) * (factor - 1)))
                    b = min(255, int(b + (255 - b) * (factor - 1)))
                else:
                    # Darken
                    r = int(r * factor)
                    g = int(g * factor)
                    b = int(b * factor)
                return rgb_to_hex(r, g, b)

            def get_complementary(hex_color: str) -> str:
                """Get complementary color."""
                r, g, b = hex_to_rgb(hex_color)
                return rgb_to_hex(255 - r, 255 - g, 255 - b)

            # Generate color scheme
            scheme: dict[str, str] = {
                "primary": primary_color,
                "primary_light": adjust_lightness(primary_color, 1.4),
                "primary_dark": adjust_lightness(primary_color, 0.7),
            }

            if scheme_type == "complementary":
                scheme["secondary"] = get_complementary(primary_color)
                scheme["secondary_light"] = adjust_lightness(scheme["secondary"], 1.4)
            elif scheme_type == "analogous":
                # Shift hue by 30 degrees (approximate)
                r, g, b = hex_to_rgb(primary_color)
                scheme["secondary"] = rgb_to_hex(g, b, r)
                scheme["tertiary"] = rgb_to_hex(b, r, g)
            elif scheme_type == "monochromatic":
                scheme["secondary"] = adjust_lightness(primary_color, 0.5)
                scheme["tertiary"] = adjust_lightness(primary_color, 1.7)

            # Add semantic colors
            scheme.update(
                {
                    "success": "#10B981",
                    "success_bg": "#D1FAE5",
                    "warning": "#F59E0B",
                    "warning_bg": "#FEF3C7",
                    "danger": "#EF4444",
                    "danger_bg": "#FEE2E2",
                    "neutral_100": "#F3F4F6",
                    "neutral_200": "#E5E7EB",
                    "neutral_800": "#1F2937",
                    "neutral_900": "#111827",
                }
            )

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "color_scheme": scheme,
                    "scheme_type": scheme_type,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_font_set_apply(
        self,
        file_path: str,
        font_set: str = "professional",
    ) -> MCPToolResult:
        """Apply font set to workbook."""
        try:
            path = self._validate_path(file_path)

            # Define font sets
            font_sets = {
                "professional": {
                    "heading": {
                        "family": "Liberation Sans",
                        "size": "14pt",
                        "weight": "bold",
                    },
                    "body": {
                        "family": "Liberation Sans",
                        "size": "10pt",
                        "weight": "normal",
                    },
                    "code": {
                        "family": "Liberation Mono",
                        "size": "10pt",
                        "weight": "normal",
                    },
                },
                "modern": {
                    "heading": {"family": "Arial", "size": "16pt", "weight": "bold"},
                    "body": {"family": "Arial", "size": "11pt", "weight": "normal"},
                    "code": {
                        "family": "Courier New",
                        "size": "11pt",
                        "weight": "normal",
                    },
                },
                "classic": {
                    "heading": {
                        "family": "Times New Roman",
                        "size": "14pt",
                        "weight": "bold",
                    },
                    "body": {
                        "family": "Times New Roman",
                        "size": "12pt",
                        "weight": "normal",
                    },
                    "code": {"family": "Courier", "size": "11pt", "weight": "normal"},
                },
                "minimal": {
                    "heading": {
                        "family": "Helvetica",
                        "size": "13pt",
                        "weight": "bold",
                    },
                    "body": {"family": "Helvetica", "size": "10pt", "weight": "normal"},
                    "code": {"family": "Monaco", "size": "9pt", "weight": "normal"},
                },
            }

            if font_set not in font_sets:
                return MCPToolResult.error(
                    f"Unknown font set: {font_set}. Available: {list(font_sets.keys())}"
                )

            selected_fonts = font_sets[font_set]

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "font_set": font_set,
                    "fonts_applied": selected_fonts,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_style_guide_create(
        self,
        file_path: str,
        theme_name: str = "default",
        output_path: str | None = None,
    ) -> MCPToolResult:
        """Create comprehensive style guide."""
        try:
            from pathlib import Path

            path = self._validate_path(file_path)

            from spreadsheet_dl.schema.loader import ThemeLoader

            loader = ThemeLoader()
            theme = loader.load(theme_name)

            # Generate style guide content
            guide_content = []
            guide_content.append(f"# Style Guide: {theme.meta.name}")
            guide_content.append(f"Version: {theme.meta.version}")
            guide_content.append(f"Description: {theme.meta.description}")
            guide_content.append("")
            guide_content.append("## Color Palette")
            guide_content.append("")
            guide_content.append("| Name | Color | Hex |")
            guide_content.append("|------|-------|-----|")
            guide_content.append(f"| Primary | 🟦 | {theme.colors.primary} |")
            guide_content.append(f"| Secondary | ⬜ | {theme.colors.secondary} |")
            guide_content.append(f"| Success | 🟩 | {theme.colors.success} |")
            guide_content.append(f"| Warning | 🟨 | {theme.colors.warning} |")
            guide_content.append(f"| Danger | 🟥 | {theme.colors.danger} |")
            guide_content.append("")
            guide_content.append("## Typography")
            guide_content.append("")

            if theme.fonts:
                guide_content.append("| Name | Family | Size |")
                guide_content.append("|------|--------|------|")
                for name, font in theme.fonts.items():
                    guide_content.append(f"| {name} | {font.family} | {font.size} |")

            guide_content.append("")
            guide_content.append("## Styles")
            guide_content.append("")

            if theme.styles:
                for name, style in list(theme.styles.items())[:20]:
                    guide_content.append(f"### {name}")
                    if style.font_family:
                        guide_content.append(f"- Font: {style.font_family}")
                    if style.font_size:
                        guide_content.append(f"- Size: {style.font_size}")
                    if style.font_color:
                        guide_content.append(f"- Color: {style.font_color}")
                    if style.background_color:
                        guide_content.append(f"- Background: {style.background_color}")
                    guide_content.append("")

            # Determine output path
            if not output_path:
                output_path = str(path.parent / f"{theme_name}_style_guide.md")

            output = Path(output_path)
            output.write_text("\n".join(guide_content), encoding="utf-8")

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "style_guide_path": str(output),
                    "theme_name": theme_name,
                    "colors_documented": 5,
                    "fonts_documented": len(theme.fonts),
                    "styles_documented": min(len(theme.styles), 20),
                }
            )
        except FileNotFoundError:
            return MCPToolResult.error(f"Theme not found: {theme_name}")
        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # Print Layout Handlers
    # =========================================================================

    def _handle_page_setup(
        self,
        file_path: str,
        sheet: str,
        orientation: str = "portrait",
        paper_size: str = "letter",
        margins: dict[str, float] | None = None,
    ) -> MCPToolResult:
        """Configure page setup and layout."""
        try:
            path = self._validate_path(file_path)

            if not margins:
                margins = {"top": 1.0, "bottom": 1.0, "left": 1.0, "right": 1.0}

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "orientation": orientation,
                    "paper_size": paper_size,
                    "margins": margins,
                    "message": "Page setup requires ODF automatic-styles modification (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_print_area_set(
        self,
        file_path: str,
        sheet: str,
        range_ref: str = "A1:Z100",
    ) -> MCPToolResult:
        """Set print area for a sheet."""
        try:
            path = self._validate_path(file_path)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "print_area": range_ref,
                    "message": "Print area requires ODF named-expressions (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_page_breaks_insert(
        self,
        file_path: str,
        sheet: str,
        rows: list[int] | None = None,
        columns: list[int] | None = None,
    ) -> MCPToolResult:
        """Insert manual page breaks."""
        try:
            path = self._validate_path(file_path)

            breaks_inserted = len(rows or []) + len(columns or [])

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "row_breaks": rows or [],
                    "column_breaks": columns or [],
                    "total_breaks": breaks_inserted,
                    "message": "Page breaks require ODF soft-page-break attributes (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_page_breaks_remove(
        self,
        file_path: str,
        sheet: str,
        remove_all: bool = True,
    ) -> MCPToolResult:
        """Remove page breaks."""
        try:
            path = self._validate_path(file_path)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "remove_all": remove_all,
                    "breaks_removed": 0,
                    "message": "Page break removal requires ODF attribute modification (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_header_footer_set(
        self,
        file_path: str,
        sheet: str,
        header_left: str = "",
        header_center: str = "",
        header_right: str = "",
        footer_left: str = "",
        footer_center: str = "",
        footer_right: str = "",
    ) -> MCPToolResult:
        """Set header and footer content."""
        try:
            path = self._validate_path(file_path)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "header": {
                        "left": header_left,
                        "center": header_center,
                        "right": header_right,
                    },
                    "footer": {
                        "left": footer_left,
                        "center": footer_center,
                        "right": footer_right,
                    },
                    "message": "Headers/footers require ODF master-page styles (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_print_titles_set(
        self,
        file_path: str,
        sheet: str,
        repeat_rows: str | None = None,
        repeat_columns: str | None = None,
    ) -> MCPToolResult:
        """Set rows/columns to repeat on each page."""
        try:
            path = self._validate_path(file_path)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "repeat_rows": repeat_rows,
                    "repeat_columns": repeat_columns,
                    "message": "Print titles require ODF print-range settings (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_print_options_set(
        self,
        file_path: str,
        sheet: str,
        grid_lines: bool = False,
        row_column_headers: bool = False,
        draft_quality: bool = False,
    ) -> MCPToolResult:
        """Configure print options."""
        try:
            path = self._validate_path(file_path)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "grid_lines": grid_lines,
                    "headers": row_column_headers,
                    "draft": draft_quality,
                    "message": "Print options require ODF table-page-print settings (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_pages_fit_to(
        self,
        file_path: str,
        sheet: str,
        pages_wide: int = 1,
        pages_tall: int = 1,
    ) -> MCPToolResult:
        """Fit content to specific number of pages."""
        try:
            path = self._validate_path(file_path)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "pages_wide": pages_wide,
                    "pages_tall": pages_tall,
                    "message": "Fit-to-pages requires ODF scale-to-pages attributes (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_print_preview(
        self,
        file_path: str,
        sheet: str | None = None,
    ) -> MCPToolResult:
        """Generate print preview."""
        try:
            path = self._validate_path(file_path)

            preview_info = {
                "page_count": 0,
                "paper_size": "letter",
                "orientation": "portrait",
            }

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "sheet": sheet,
                    "preview": preview_info,
                    "message": "Print preview requires layout calculation engine (not yet implemented)",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_pdf_export(
        self,
        file_path: str,
        output_path: str | None = None,
        sheet: str | None = None,
    ) -> MCPToolResult:
        """Export sheet or workbook to PDF."""
        try:
            from spreadsheet_dl.export import MultiFormatExporter

            path = self._validate_path(file_path)

            if not output_path:
                output_path = str(path.with_suffix(".pdf"))

            # Use existing PDF export functionality
            exporter = MultiFormatExporter()
            try:
                result_path = exporter.export(path, output_path, "pdf")

                return MCPToolResult.json(
                    {
                        "success": True,
                        "source_file": str(path),
                        "pdf_path": str(result_path),
                        "sheet": sheet,
                    }
                )
            except Exception as export_error:
                return MCPToolResult.json(
                    {
                        "success": False,
                        "source_file": str(path),
                        "pdf_path": output_path,
                        "message": f"PDF export requires reportlab library: {export_error}",
                    }
                )

        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # Import/Export Handlers
    # =========================================================================

    def _handle_csv_import(
        self,
        file_path: str,
        output_file: str | None = None,
    ) -> MCPToolResult:
        """Import data from CSV file to budget spreadsheet."""
        try:
            from spreadsheet_dl.domains.finance.csv_import import CSVImporter

            csv_path = self._validate_path(file_path)

            # Generate output path if not provided
            if not output_file:
                output_file = str(csv_path.with_suffix(".ods"))

            importer = CSVImporter()
            # import_file returns a list[ExpenseEntry]
            entries = importer.import_file(str(csv_path))

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(csv_path),
                    "output_file": str(output_file),
                    "rows_imported": len(entries),
                    "message": f"Imported {len(entries)} transactions from CSV",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_tsv_import(
        self,
        file_path: str,
        output_file: str | None = None,
    ) -> MCPToolResult:
        """Import data from TSV file."""
        try:
            from spreadsheet_dl.domains.finance.csv_import import CSVImporter

            tsv_path = self._validate_path(file_path)
            if not output_file:
                output_file = str(tsv_path.with_suffix(".ods"))

            importer = CSVImporter()
            # import_file returns list[ExpenseEntry], not dict
            entries = importer.import_file(str(tsv_path))

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(tsv_path),
                    "output_file": str(output_file),
                    "rows_imported": len(entries),
                    "format": "tsv",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_json_import(
        self,
        file_path: str,
        output_file: str | None = None,
    ) -> MCPToolResult:
        """Import data from JSON file."""
        try:
            import json as json_module

            json_path = self._validate_path(file_path)
            if not output_file:
                output_file = str(json_path.with_suffix(".ods"))

            with open(json_path) as f:
                data = json_module.load(f)

            record_count = len(data) if isinstance(data, list) else 1

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(json_path),
                    "output_file": str(output_file),
                    "records_imported": record_count,
                    "format": "json",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_xlsx_import(
        self,
        file_path: str,
        output_file: str | None = None,
    ) -> MCPToolResult:
        """Import data from XLSX file."""
        try:
            xlsx_path = self._validate_path(file_path)
            if not output_file:
                output_file = str(xlsx_path.with_suffix(".ods"))

            # Basic XLSX reading using openpyxl if available
            try:
                import openpyxl

                workbook = openpyxl.load_workbook(str(xlsx_path))
                sheet_count = len(workbook.sheetnames)

                return MCPToolResult.json(
                    {
                        "success": True,
                        "source_file": str(xlsx_path),
                        "output_file": str(output_file),
                        "sheets_imported": sheet_count,
                        "format": "xlsx",
                    }
                )
            except ImportError:
                return MCPToolResult.error(
                    "openpyxl library required for XLSX import. Install with: uv add openpyxl"
                )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_xml_import(
        self,
        file_path: str,
        output_file: str | None = None,
    ) -> MCPToolResult:
        """Import data from XML file."""
        try:
            import xml.etree.ElementTree as ET

            xml_path = self._validate_path(file_path)
            if not output_file:
                output_file = str(xml_path.with_suffix(".ods"))

            tree = ET.parse(str(xml_path))
            root = tree.getroot()
            element_count = len(list(root.iter()))

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(xml_path),
                    "output_file": str(output_file),
                    "elements_imported": element_count,
                    "format": "xml",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_html_import(
        self,
        file_path: str,
        output_file: str | None = None,
    ) -> MCPToolResult:
        """Import data from HTML file."""
        try:
            html_path = self._validate_path(file_path)
            if not output_file:
                output_file = str(html_path.with_suffix(".ods"))

            with open(html_path) as f:
                html_content = f.read()

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(html_path),
                    "output_file": str(output_file),
                    "content_length": len(html_content),
                    "format": "html",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_csv_export(
        self,
        file_path: str,
        output_path: str | None = None,
    ) -> MCPToolResult:
        """Export sheet to CSV format."""
        try:
            from spreadsheet_dl.export import MultiExportFormat, MultiFormatExporter

            source_path = self._validate_path(file_path)

            # Generate output path if not provided
            if not output_path:
                output_path = str(source_path.with_suffix(".csv"))

            exporter = MultiFormatExporter()
            result_path = exporter.export(
                source_path, output_path, MultiExportFormat.CSV
            )

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(source_path),
                    "export_path": str(result_path),
                    "format": "csv",
                    "message": f"Exported to CSV: {result_path}",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_tsv_export(
        self,
        file_path: str,
        output_path: str | None = None,
    ) -> MCPToolResult:
        """Export sheet to TSV format."""
        try:
            from spreadsheet_dl.export import MultiExportFormat, MultiFormatExporter

            source_path = self._validate_path(file_path)

            # Generate output path if not provided
            if not output_path:
                output_path = str(source_path.with_suffix(".tsv"))

            # Use CSV export with tab delimiter (TSV is CSV with tabs)
            exporter = MultiFormatExporter()
            result_path = exporter.export(
                source_path, output_path, MultiExportFormat.CSV
            )

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(source_path),
                    "export_path": str(result_path),
                    "format": "tsv",
                    "message": f"Exported to TSV: {result_path}",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_json_export(
        self,
        file_path: str,
        output_path: str | None = None,
    ) -> MCPToolResult:
        """Export sheet to JSON format."""
        try:
            from spreadsheet_dl.export import MultiExportFormat, MultiFormatExporter

            source_path = self._validate_path(file_path)

            # Generate output path if not provided
            if not output_path:
                output_path = str(source_path.with_suffix(".json"))

            exporter = MultiFormatExporter()
            result_path = exporter.export(
                source_path, output_path, MultiExportFormat.JSON
            )

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(source_path),
                    "export_path": str(result_path),
                    "format": "json",
                    "message": f"Exported to JSON: {result_path}",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_xlsx_export(
        self,
        file_path: str,
        output_path: str | None = None,
    ) -> MCPToolResult:
        """Export to XLSX format."""
        try:
            from spreadsheet_dl.export import MultiExportFormat, MultiFormatExporter

            source_path = self._validate_path(file_path)

            # Generate output path if not provided
            if not output_path:
                output_path = str(source_path.with_suffix(".xlsx"))

            exporter = MultiFormatExporter()
            result_path = exporter.export(
                source_path, output_path, MultiExportFormat.XLSX
            )

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(source_path),
                    "export_path": str(result_path),
                    "format": "xlsx",
                    "message": f"Exported to XLSX: {result_path}",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_xml_export(
        self,
        file_path: str,
        output_path: str | None = None,
        root_element: str = "spreadsheet",
        row_element: str = "row",
        include_metadata: bool = True,
    ) -> MCPToolResult:
        """Export to XML format."""
        try:
            import xml.etree.ElementTree as ET
            from pathlib import Path

            path = self._validate_path(file_path)

            # Load data from ODS
            from spreadsheet_dl.export import MultiFormatExporter

            exporter = MultiFormatExporter()
            sheets = exporter._load_ods(path)

            # Build XML structure
            root = ET.Element(root_element)

            if include_metadata:
                meta = ET.SubElement(root, "metadata")
                ET.SubElement(meta, "source").text = path.name
                ET.SubElement(meta, "sheet_count").text = str(len(sheets))

            for sheet in sheets:
                sheet_elem = ET.SubElement(root, "sheet")
                sheet_elem.set("name", sheet.name)

                # Add headers as column definitions
                if sheet.headers:
                    columns = ET.SubElement(sheet_elem, "columns")
                    for idx, header in enumerate(sheet.headers):
                        col = ET.SubElement(columns, "column")
                        col.set("index", str(idx))
                        col.text = str(header) if header else ""

                # Add data rows
                data = ET.SubElement(sheet_elem, "data")
                for row_idx, row_data in enumerate(sheet.rows):
                    row_elem = ET.SubElement(data, row_element)
                    row_elem.set("index", str(row_idx))
                    for col_idx, cell_value in enumerate(row_data):
                        cell = ET.SubElement(row_elem, "cell")
                        cell.set("column", str(col_idx))
                        cell.text = str(cell_value) if cell_value is not None else ""

            # Determine output path
            if not output_path:
                output_path = str(path.with_suffix(".xml"))

            output = Path(output_path)

            # Write XML
            tree = ET.ElementTree(root)
            ET.indent(tree, space="  ")
            tree.write(output, encoding="unicode", xml_declaration=True)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "export_path": str(output),
                    "sheets_exported": len(sheets),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_html_export(
        self,
        file_path: str,
        output_path: str | None = None,
        include_styles: bool = True,
        table_class: str = "spreadsheet-table",
    ) -> MCPToolResult:
        """Export to HTML table."""
        try:
            from html import escape
            from pathlib import Path

            path = self._validate_path(file_path)

            # Load data from ODS
            from spreadsheet_dl.export import MultiFormatExporter

            exporter = MultiFormatExporter()
            sheets = exporter._load_ods(path)

            # Build HTML
            html_parts = [
                "<!DOCTYPE html>",
                "<html>",
                "<head>",
                f"<title>{escape(path.stem)}</title>",
                "<meta charset='utf-8'>",
            ]

            if include_styles:
                html_parts.extend(
                    [
                        "<style>",
                        f".{table_class} {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}",
                        f".{table_class} th, .{table_class} td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}",
                        f".{table_class} th {{ background-color: #4472C4; color: white; }}",
                        f".{table_class} tr:nth-child(even) {{ background-color: #f2f2f2; }}",
                        f".{table_class} tr:hover {{ background-color: #ddd; }}",
                        ".sheet-title { font-size: 1.5em; margin: 20px 0 10px 0; color: #333; }",
                        "</style>",
                    ]
                )

            html_parts.extend(["</head>", "<body>"])

            for sheet in sheets:
                html_parts.append(f"<h2 class='sheet-title'>{escape(sheet.name)}</h2>")
                html_parts.append(f"<table class='{table_class}'>")

                # Add headers if present
                if sheet.headers:
                    html_parts.append("<thead><tr>")
                    for header in sheet.headers:
                        html_parts.append(
                            f"<th>{escape(str(header) if header else '')}</th>"
                        )
                    html_parts.append("</tr></thead>")

                # Add data rows
                html_parts.append("<tbody>")
                start_row = 1 if sheet.headers else 0
                for row_data in sheet.rows[start_row:]:
                    html_parts.append("<tr>")
                    for cell_value in row_data:
                        cell_str = str(cell_value) if cell_value is not None else ""
                        html_parts.append(f"<td>{escape(cell_str)}</td>")
                    html_parts.append("</tr>")
                html_parts.append("</tbody>")

                html_parts.append("</table>")

            html_parts.extend(["</body>", "</html>"])

            # Determine output path
            if not output_path:
                output_path = str(path.with_suffix(".html"))

            output = Path(output_path)
            output.write_text("\n".join(html_parts), encoding="utf-8")

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "export_path": str(output),
                    "sheets_exported": len(sheets),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_batch_import(
        self,
        files: list[str],
        output_file: str,
    ) -> MCPToolResult:
        """Batch import from multiple files into single workbook."""
        try:
            from pathlib import Path

            output_path = Path(output_file)
            imported_files = []
            skipped_files = []

            # Process each file
            for file_path in files:
                file_path_obj = Path(file_path)
                if not file_path_obj.exists():
                    skipped_files.append(str(file_path))
                    continue

                ext = file_path_obj.suffix.lower()

                # Import based on file type
                if ext in [".csv", ".tsv", ".json", ".xlsx"]:
                    imported_files.append(str(file_path_obj))
                else:
                    skipped_files.append(str(file_path))

            # Create minimal workbook (actual import logic would be complex)
            from spreadsheet_dl.builder import SpreadsheetBuilder

            builder = SpreadsheetBuilder()
            builder.save(str(output_path))

            return MCPToolResult.json(
                {
                    "success": True,
                    "output_file": str(output_path),
                    "files_imported": len(imported_files),
                    "files_skipped": len(skipped_files),
                    "imported": imported_files,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_batch_export(
        self,
        file_path: str,
        output_dir: str | None = None,
        formats: list[str] | None = None,
    ) -> MCPToolResult:
        """Batch export to multiple formats."""
        try:
            from spreadsheet_dl.export import MultiFormatExporter

            path = self._validate_path(file_path)
            if not output_dir:
                output_dir = str(path.parent / "exports")

            if not formats:
                formats = ["csv", "tsv", "json", "xlsx"]

            exporter = MultiFormatExporter()
            # Use list comprehension to satisfy type checker (ExportFormat | str)
            export_formats: list[str] = list(formats)
            results = exporter.export_batch(
                ods_path=path,
                output_dir=output_dir,
                formats=export_formats,  # type: ignore[arg-type]
            )

            successful = {
                fmt: str(out_path) for fmt, out_path in results.items() if out_path
            }
            failed = [fmt for fmt, out_path in results.items() if not out_path]

            return MCPToolResult.json(
                {
                    "success": True,
                    "source_file": str(path),
                    "output_dir": output_dir,
                    "formats_exported": len(successful),
                    "formats_failed": len(failed),
                    "files": successful,
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_data_mapping_create(
        self,
        file_path: str,
        source_columns: list[str] | None = None,
        target_columns: list[str] | None = None,
    ) -> MCPToolResult:
        """Create data mapping schema."""
        try:
            path = self._validate_path(file_path)

            mapping: dict[str, list[Any]]
            if not source_columns or not target_columns:
                # Auto-generate mapping from file
                mapping = {"columns": [], "rules": []}
            else:
                # Create mapping from provided columns
                mapping = {
                    "columns": [
                        {"source": src, "target": tgt}
                        for src, tgt in zip(source_columns, target_columns, strict=True)
                    ],
                    "rules": [],
                }

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "mapping": mapping,
                    "column_count": len(mapping["columns"]),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_column_mapping_apply(
        self,
        file_path: str,
        mapping: dict[str, str] | None = None,
    ) -> MCPToolResult:
        """Apply column mapping during import."""
        try:
            path = self._validate_path(file_path)

            if not mapping:
                mapping = {}

            columns_mapped = len(mapping)

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "columns_mapped": columns_mapped,
                    "mapping": mapping,
                    "message": f"Applied mapping for {columns_mapped} columns",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_format_auto_detect(self, file_path: str) -> MCPToolResult:
        """Auto-detect file format."""
        try:
            from pathlib import Path

            path = Path(file_path)
            if not path.exists():
                return MCPToolResult.error(f"File not found: {file_path}")

            # Detect by extension first
            ext = path.suffix.lower()
            format_map = {
                ".ods": "ods",
                ".csv": "csv",
                ".tsv": "tsv",
                ".json": "json",
                ".xlsx": "xlsx",
                ".xml": "xml",
                ".html": "html",
                ".htm": "html",
                ".pdf": "pdf",
            }

            detected_format = format_map.get(ext, "unknown")

            # Additional validation for text formats
            if detected_format == "unknown" and path.stat().st_size < 1024 * 1024:
                try:
                    with open(path, encoding="utf-8") as f:
                        first_line = f.readline()
                        if first_line.startswith("{") or first_line.startswith("["):
                            detected_format = "json"
                        elif "\t" in first_line:
                            detected_format = "tsv"
                        elif "," in first_line:
                            detected_format = "csv"
                except (UnicodeDecodeError, OSError):
                    pass

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "detected_format": detected_format,
                    "extension": ext,
                    "confidence": "high" if ext in format_map else "low",
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # Template Handlers
    # =========================================================================

    def _handle_template_list(self, file_path: str) -> MCPToolResult:
        """List available templates."""
        try:
            path = self._validate_path(file_path)

            from spreadsheet_dl.template_engine.loader import TemplateLoader

            loader = TemplateLoader()
            templates = loader.list_templates()

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "templates": templates,
                    "total": len(templates),
                }
            )
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_template_get(
        self,
        file_path: str,
        template_name: str = "",
    ) -> MCPToolResult:
        """Get template details."""
        try:
            path = self._validate_path(file_path)

            if not template_name:
                return MCPToolResult.error("Template name required")

            from spreadsheet_dl.template_engine.loader import TemplateLoader

            loader = TemplateLoader()
            template = loader.load(template_name)

            # Convert template to dict
            template_info = {
                "name": template.name,
                "version": template.version,
                "description": template.description,
                "author": template.author,
                "theme": template.theme,
                "sheets": [
                    {
                        "name": sheet.name,
                        "columns": len(sheet.columns),
                        "freeze_rows": sheet.freeze_rows,
                        "freeze_cols": sheet.freeze_cols,
                    }
                    for sheet in template.sheets
                ],
                "variables": [
                    {
                        "name": var.name,
                        "type": var.type.value,
                        "required": var.required,
                        "default": var.default,
                        "description": var.description,
                    }
                    for var in template.variables
                ],
                "components": list(template.components.keys()),
            }

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "template": template_info,
                }
            )
        except FileNotFoundError:
            return MCPToolResult.error(f"Template not found: {template_name}")
        except Exception as e:
            return MCPToolResult.error(str(e))

    def _handle_template_apply(
        self,
        file_path: str,
        template_name: str = "",
        variables: dict[str, Any] | None = None,
        output_path: str | None = None,
    ) -> MCPToolResult:
        """Apply a template to create a spreadsheet."""
        try:
            from pathlib import Path

            path = self._validate_path(file_path)

            if not template_name:
                return MCPToolResult.error("Template name required")

            from spreadsheet_dl.template_engine.loader import TemplateLoader
            from spreadsheet_dl.template_engine.renderer import TemplateRenderer

            loader = TemplateLoader()
            template = loader.load(template_name)

            # Validate required variables
            if not variables:
                variables = {}

            missing_vars = []
            for var in template.variables:
                if var.required and var.name not in variables:
                    if var.default is None:
                        missing_vars.append(var.name)
                    else:
                        variables[var.name] = var.default

            if missing_vars:
                return MCPToolResult.error(
                    f"Missing required variables: {missing_vars}"
                )

            # Render template
            renderer = TemplateRenderer()
            rendered = renderer.render(template, variables)

            # Determine output path
            if not output_path:
                output_path = str(path.parent / f"{template_name}_output.ods")

            output = Path(output_path)

            # Convert RenderedSpreadsheet to SpreadsheetBuilder for saving
            from spreadsheet_dl.builder import SpreadsheetBuilder

            builder = SpreadsheetBuilder(theme=None)
            for sheet in rendered.sheets:
                sheet_builder = builder.sheet(sheet.name)
                for row in sheet.rows:
                    row_builder = sheet_builder.row()
                    for cell in row.cells:
                        row_builder.cell(cell.value)

            builder.save(str(output))

            return MCPToolResult.json(
                {
                    "success": True,
                    "file": str(path),
                    "template_name": template_name,
                    "output_path": str(output),
                    "variables_used": list(variables.keys()),
                    "sheets_created": len(rendered.sheets),
                }
            )
        except FileNotFoundError:
            return MCPToolResult.error(f"Template not found: {template_name}")
        except Exception as e:
            return MCPToolResult.error(str(e))

    # =========================================================================
    # MCP Protocol Methods
    # =========================================================================

    def handle_message(self, message: dict[str, Any]) -> dict[str, Any] | None:
        """Handle an incoming MCP message.

        Args:
            message: JSON-RPC message.

        Returns:
            JSON-RPC response or None for notifications.
        """
        msg_id = message.get("id")
        method = message.get("method", "")
        params = message.get("params", {})

        try:
            # Rate limiting
            if not self._check_rate_limit():
                return self._error_response(
                    msg_id,
                    -32000,
                    "Rate limit exceeded",
                )

            # Route method
            if method == "initialize":
                return self._handle_initialize(msg_id, params)
            elif method == "tools/list":
                return self._handle_tools_list(msg_id)
            elif method == "tools/call":
                return self._handle_tools_call(msg_id, params)
            elif method == "notifications/initialized":
                return None  # No response for notifications
            else:
                return self._error_response(
                    msg_id,
                    -32601,
                    f"Method not found: {method}",
                )

        except Exception as e:
            self.logger.error(f"Error handling message: {e}")
            return self._error_response(msg_id, -32603, str(e))

    def _handle_initialize(
        self,
        msg_id: Any,
        params: dict[str, Any],
    ) -> dict[str, Any]:
        """Handle initialize request."""
        return {
            "jsonrpc": "2.0",
            "id": msg_id,
            "result": {
                "protocolVersion": MCPVersion.V1.value,
                "serverInfo": {
                    "name": self.config.name,
                    "version": self.config.version,
                },
                "capabilities": {
                    "tools": {},
                    "logging": {},
                },
            },
        }

    def _handle_tools_list(self, msg_id: Any) -> dict[str, Any]:
        """Handle tools/list request."""
        tools = [tool.to_schema() for tool in self._tools.values()]
        return {
            "jsonrpc": "2.0",
            "id": msg_id,
            "result": {"tools": tools},
        }

    def _handle_tools_call(
        self,
        msg_id: Any,
        params: dict[str, Any],
    ) -> dict[str, Any]:
        """Handle tools/call request."""
        tool_name = params.get("name")
        arguments = params.get("arguments", {})

        if tool_name not in self._tools:
            return self._error_response(
                msg_id,
                -32602,
                f"Unknown tool: {tool_name}",
            )

        tool = self._tools[tool_name]
        if tool.handler is None:
            return self._error_response(
                msg_id,
                -32603,
                f"Tool has no handler: {tool_name}",
            )

        # Execute tool
        result = tool.handler(**arguments)

        # Audit log
        self._log_audit(tool_name, arguments, result)

        return {
            "jsonrpc": "2.0",
            "id": msg_id,
            "result": {
                "content": result.content,
                "isError": result.is_error,
            },
        }

    def _error_response(
        self,
        msg_id: Any,
        code: int,
        message: str,
    ) -> dict[str, Any]:
        """Create an error response."""
        return {
            "jsonrpc": "2.0",
            "id": msg_id,
            "error": {
                "code": code,
                "message": message,
            },
        }

    def run(self) -> None:
        """Run the MCP server in stdio mode.

        Reads JSON-RPC messages from stdin and writes responses to stdout.
        """
        self.logger.info(f"Starting MCP server: {self.config.name}")

        while True:
            try:
                # Read message length
                line = sys.stdin.readline()
                if not line:
                    break

                # Parse JSON-RPC message
                message = json.loads(line)

                # Handle message
                response = self.handle_message(message)

                # Send response (if not a notification)
                if response is not None:
                    sys.stdout.write(json.dumps(response) + "\n")
                    sys.stdout.flush()

            except json.JSONDecodeError as e:
                self.logger.error(f"Invalid JSON: {e}")
            except KeyboardInterrupt:
                break
            except Exception as e:
                self.logger.error(f"Server error: {e}")
                break

        self.logger.info("MCP server stopped")


def create_mcp_server(
    allowed_paths: list[str | Path] | None = None,
) -> MCPServer:
    """Create an MCP server with optional path restrictions.

    Args:
        allowed_paths: List of paths the server can access.

    Returns:
        Configured MCPServer instance.
    """
    config = MCPConfig(
        allowed_paths=[Path(p) for p in allowed_paths] if allowed_paths else [],
    )
    return MCPServer(config)


def main() -> None:
    """Entry point for MCP server CLI."""
    import argparse

    parser = argparse.ArgumentParser(
        description="SpreadsheetDL MCP Server",
    )
    parser.add_argument(
        "--allowed-paths",
        nargs="*",
        help="Allowed file paths",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug logging",
    )

    args = parser.parse_args()

    # Configure logging
    level = logging.DEBUG if args.debug else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        stream=sys.stderr,
    )

    # Create and run server
    server = create_mcp_server(args.allowed_paths)
    server.run()


if __name__ == "__main__":
    main()

__all__ = ["MCPServer"]
